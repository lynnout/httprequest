#-*- coding:utf-8 -*-
# @Time  : 2019/4/17 18:15
# @Author: lisa.liu
# @Email : 2423844@qq.com
# @File  : DoExcel.py

from openpyxl import load_workbook
from common import config

class DoExcel:
    def __init__(self,filename,sheetname):
        self.filename=filename
        self.workbook=load_workbook(filename)
        self.sheet=self.workbook[sheetname]

    #读excel操作
    def get_cases(self):
        max_row=self.sheet.max_row
        cases = []
        for i in range(2,max_row+1):
             case=Case()
             case.case_id = self.sheet.cell(row=i,column=1).value
             case.title = self.sheet.cell(row=i,column=2).value
             case.url = url=config.ReadConfig().get('api','pre_url')+self.sheet.cell(row=i, column=3).value
             case.data = self.sheet.cell(row=i, column=4).value
             case.method = self.sheet.cell(row=i, column=5).value
             case.expected = self.sheet.cell(row=i, column=6).value
             case.sql = self.sheet.cell(row=i,column=9).value
             cases.append(case)

        self.workbook.close()
        return cases

    #写excel操作
    def write_result(self,row,actual,result):
        self.sheet.cell(row,7).value=actual
        self.sheet.cell(row,8).value=result
        self.workbook.save(filename=self.filename)
        self.workbook.close()



class Case:
    def __init__(self):
        self.case_id=None
        self.title=None
        self.url = None
        self.data = None
        self.method=None
        self.expected = None
        self.actual = None
        self.result = None

if __name__=='__main__':
    myexcel=DoExcel('../data/观观cases.xlsx','register')
    cases=myexcel.get_cases()
    for case in cases:
        print(case.case_id,case.title,case.url,case.data,case.method,case.expected)